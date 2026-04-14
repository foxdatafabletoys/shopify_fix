#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
import psycopg


EXPECTED_HEADERS = [
    "Brand",
    "TITLE",
    "EAN",
    "SKU",
    "ASIN",
    "RRP (excl vat) £",
    "RRP (excl vat) $",
    "Buy cost exc. VAT £",
    "Markup",
    "Trade Price ex VAT £",
    "Trade Price ex VAT $",
    "Discount vs RRP",
    "Case Pack",
    "UK",
    "US",
]


@dataclass
class ProductRow:
    brand: str
    name: str
    sku: str
    barcode: str
    asin: str
    category: str
    rrp_ex_vat_gbp: Decimal | None
    rrp_ex_vat_usd: Decimal | None
    buy_cost_exc_vat_gbp: Decimal
    trade_price_ex_vat_gbp: Decimal
    trade_price_ex_vat_usd: Decimal | None
    quoted_usd_per_gbp: Decimal | None
    markup: Decimal | None
    discount_vs_rrp: Decimal | None
    case_pack: int | None
    cost_price_gbp: Decimal
    sell_price_gbp: Decimal
    uk_flag: str
    us_flag: str


def as_trimmed_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_money(value: object) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def as_optional_money(value: object) -> Decimal | None:
    if value is None or as_trimmed_text(value) == "":
        return None
    return as_money(value)


def as_optional_decimal_4(value: object) -> Decimal | None:
    if value is None or as_trimmed_text(value) == "":
        return None
    return Decimal(str(value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def as_optional_decimal_6(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value)).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)


def as_optional_int(value: object) -> int | None:
    text_value = as_trimmed_text(value)
    if text_value == "":
        return None
    return int(float(text_value))


def normalize_flag(value: object) -> str:
    return as_trimmed_text(value).upper()


def load_products(xlsx_path: Path) -> list[ProductRow]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active

    header = [as_trimmed_text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[:15]]
    if header != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected header row in {xlsx_path}. Found: {header}")

    products: list[ProductRow] = []
    seen_skus: set[str] = set()

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(as_trimmed_text(cell) for cell in excel_row[:15]):
            continue

        name = as_trimmed_text(excel_row[1])
        sku = as_trimmed_text(excel_row[3])
        barcode = as_trimmed_text(excel_row[2])
        brand = as_trimmed_text(excel_row[0])
        asin = as_trimmed_text(excel_row[4])
        category = brand
        rrp_ex_vat_gbp = as_optional_money(excel_row[5])
        rrp_ex_vat_usd = as_optional_money(excel_row[6])
        buy_cost_exc_vat_gbp = as_money(excel_row[7])
        markup = as_optional_decimal_4(excel_row[8])
        trade_price_ex_vat_gbp = as_money(excel_row[9])
        trade_price_ex_vat_usd = as_optional_money(excel_row[10])
        discount_vs_rrp = as_optional_decimal_4(excel_row[11])
        case_pack = as_optional_int(excel_row[12])
        uk_flag = normalize_flag(excel_row[13])
        us_flag = normalize_flag(excel_row[14])
        cost_price_gbp = buy_cost_exc_vat_gbp
        sell_price_gbp = trade_price_ex_vat_gbp
        quoted_usd_per_gbp = None
        if trade_price_ex_vat_usd is not None and trade_price_ex_vat_gbp > 0:
            quoted_usd_per_gbp = as_optional_decimal_6(trade_price_ex_vat_usd / trade_price_ex_vat_gbp)
        elif rrp_ex_vat_usd is not None and rrp_ex_vat_gbp is not None and rrp_ex_vat_gbp > 0:
            quoted_usd_per_gbp = as_optional_decimal_6(rrp_ex_vat_usd / rrp_ex_vat_gbp)

        if not name:
            raise ValueError("Found row with empty TITLE.")
        if not sku:
            raise ValueError("Found row with empty SKU.")
        if sku in seen_skus:
            raise ValueError(f"Duplicate SKU in spreadsheet: {sku}")
        if uk_flag not in {"Y", "N"} or us_flag not in {"Y", "N"}:
            raise ValueError(f"Invalid UK/US flag for SKU {sku}: UK={uk_flag!r}, US={us_flag!r}")

        seen_skus.add(sku)
        products.append(
            ProductRow(
                brand=brand,
                name=name,
                sku=sku,
                barcode=barcode,
                asin=asin,
                category=category,
                rrp_ex_vat_gbp=rrp_ex_vat_gbp,
                rrp_ex_vat_usd=rrp_ex_vat_usd,
                buy_cost_exc_vat_gbp=buy_cost_exc_vat_gbp,
                trade_price_ex_vat_gbp=trade_price_ex_vat_gbp,
                trade_price_ex_vat_usd=trade_price_ex_vat_usd,
                quoted_usd_per_gbp=quoted_usd_per_gbp,
                markup=markup,
                discount_vs_rrp=discount_vs_rrp,
                case_pack=case_pack,
                cost_price_gbp=cost_price_gbp,
                sell_price_gbp=sell_price_gbp,
                uk_flag=uk_flag,
                us_flag=us_flag,
            )
        )

    return products


def reseed_inventory(dsn: str, products: list[ProductRow]) -> None:
    with psycopg.connect(dsn, sslmode="require") as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                create extension if not exists pgcrypto;

                create table if not exists products (
                  id uuid primary key default gen_random_uuid(),
                  brand text,
                  name text not null,
                  sku text unique not null,
                  barcode text,
                  asin text,
                  rrp_ex_vat_gbp numeric(10,2),
                  rrp_ex_vat_usd numeric(10,2),
                  buy_cost_exc_vat_gbp numeric(10,2),
                  trade_price_ex_vat_gbp numeric(10,2),
                  trade_price_ex_vat_usd numeric(10,2),
                  quoted_usd_per_gbp numeric(12,6),
                  fx_quote_date date,
                  markup numeric(10,4),
                  discount_vs_rrp numeric(10,4),
                  case_pack integer,
                  uk_flag text,
                  us_flag text,
                  category text,
                  cost_price_gbp numeric(10,2) not null,
                  sell_price_gbp numeric(10,2) not null,
                  weight_kg numeric(6,3),
                  created_at timestamptz default now()
                );

                create table if not exists warehouse_stock (
                  id uuid primary key default gen_random_uuid(),
                  product_id uuid references products(id) on delete cascade,
                  warehouse text not null,
                  qty_on_hand integer not null default 0,
                  qty_reserved integer not null default 0,
                  reorder_point integer not null default 10,
                  unique(product_id, warehouse)
                );
                """
            )

            cursor.execute("alter table products add column if not exists brand text;")
            cursor.execute("alter table products add column if not exists asin text;")
            cursor.execute("alter table products add column if not exists rrp_ex_vat_gbp numeric(10,2);")
            cursor.execute("alter table products add column if not exists rrp_ex_vat_usd numeric(10,2);")
            cursor.execute("alter table products add column if not exists buy_cost_exc_vat_gbp numeric(10,2);")
            cursor.execute("alter table products add column if not exists trade_price_ex_vat_gbp numeric(10,2);")
            cursor.execute("alter table products add column if not exists trade_price_ex_vat_usd numeric(10,2);")
            cursor.execute("alter table products add column if not exists quoted_usd_per_gbp numeric(12,6);")
            cursor.execute("alter table products add column if not exists fx_quote_date date;")
            cursor.execute("alter table products add column if not exists markup numeric(10,4);")
            cursor.execute("alter table products add column if not exists discount_vs_rrp numeric(10,4);")
            cursor.execute("alter table products add column if not exists case_pack integer;")
            cursor.execute("alter table products add column if not exists uk_flag text;")
            cursor.execute("alter table products add column if not exists us_flag text;")

            cursor.execute("alter table warehouse_stock drop constraint if exists warehouse_stock_warehouse_check;")
            cursor.execute("truncate table warehouse_stock, products;")
            cursor.execute(
                "alter table warehouse_stock add constraint warehouse_stock_warehouse_check check (warehouse in ('uk','us'));"
            )

            for product in products:
                cursor.execute(
                    """
                    insert into products (
                      brand, name, sku, barcode, asin, rrp_ex_vat_gbp, rrp_ex_vat_usd, buy_cost_exc_vat_gbp,
                      trade_price_ex_vat_gbp, trade_price_ex_vat_usd, quoted_usd_per_gbp, fx_quote_date,
                      markup, discount_vs_rrp, case_pack, uk_flag, us_flag,
                      category, cost_price_gbp, sell_price_gbp
                    ) values (
                      %s, %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s, %s, %s,
                      %s, %s, %s
                    )
                    returning id;
                    """,
                    (
                        product.brand or None,
                        product.name,
                        product.sku,
                        product.barcode or None,
                        product.asin or None,
                        product.rrp_ex_vat_gbp,
                        product.rrp_ex_vat_usd,
                        product.buy_cost_exc_vat_gbp,
                        product.trade_price_ex_vat_gbp,
                        product.trade_price_ex_vat_usd,
                        product.quoted_usd_per_gbp,
                        None,
                        product.markup,
                        product.discount_vs_rrp,
                        product.case_pack,
                        product.uk_flag,
                        product.us_flag,
                        product.category or None,
                        product.cost_price_gbp,
                        product.sell_price_gbp,
                    ),
                )
                product_id = cursor.fetchone()[0]

                stock_rows = [
                    ("uk", 1 if product.uk_flag == "Y" else 0),
                    ("us", 1 if product.us_flag == "Y" else 0),
                ]
                for warehouse, qty_on_hand in stock_rows:
                    cursor.execute(
                        """
                        insert into warehouse_stock (
                          product_id, warehouse, qty_on_hand, qty_reserved, reorder_point
                        ) values (%s, %s, %s, 0, 0);
                        """,
                        (product_id, warehouse, qty_on_hand),
                    )

        connection.commit()


def print_verification(dsn: str) -> None:
    with psycopg.connect(dsn, sslmode="require") as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                select
                  (select count(*) from products) as products_count,
                  (select count(*) from warehouse_stock) as stock_rows_count,
                  (select count(*) from warehouse_stock where warehouse = 'uk' and qty_on_hand > 0) as uk_y_count,
                  (select count(*) from warehouse_stock where warehouse = 'us' and qty_on_hand > 0) as us_y_count
                """
            )
            row = cursor.fetchone()
            print(
                f"products={row[0]} warehouse_stock={row[1]} uk_y={row[2]} us_y={row[3]}"
            )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reseed Telemachus inventory in Supabase from stock XLSX."
    )
    parser.add_argument(
        "--file",
        required=True,
        help="Absolute path to stock XLSX file.",
    )
    parser.add_argument(
        "--dsn",
        default=os.environ.get("SUPABASE_DSN", ""),
        help="PostgreSQL DSN (or set SUPABASE_DSN).",
    )
    args = parser.parse_args()

    dsn = args.dsn.strip()
    if not dsn:
        raise ValueError("Missing DSN. Provide --dsn or SUPABASE_DSN.")

    xlsx_path = Path(args.file).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Stock file not found: {xlsx_path}")

    products = load_products(xlsx_path)
    reseed_inventory(dsn, products)
    print_verification(dsn)


if __name__ == "__main__":
    main()
